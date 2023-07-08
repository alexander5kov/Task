import requests
import pandas as pd
import argparse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from io import BytesIO

parser = argparse.ArgumentParser()
parser.add_argument('-k', '--keys', nargs='+', help='Additional keys for columns')
parser.add_argument('-c', '--colored', action='store_true', help='Enable row coloring', default=True)
args = parser.parse_args()

csv_file = 'vehicles.csv'
df = pd.read_csv(csv_file, delimiter=';')

# Login to retrieve access token
login_url = "https://api.baubuddy.de/index.php/login"
login_payload = {
    "username": "365",
    "password": "1"
}
login_headers = {
    "Authorization": "Basic QVBJX0V4cGxvcmVyOjEyMzQ1NmlzQUxhbWVQYXNz",
    "Content-Type": "application/json"
}
login_response = requests.post(login_url, json=login_payload, headers=login_headers)

if login_response.status_code == 200:
    access_token = login_response.json()["oauth"]["access_token"]

    url = 'https://api.baubuddy.de/dev/index.php/v1/vehicles/select/active'

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    payload = df.to_json(orient='records')

    response = requests.post(url, json=payload, headers=headers)

    if response.status_code == 200:
        data = response.json()
        df = pd.DataFrame(data)
        df.sort_values(by='gruppe', inplace=True)
        df.insert(0, 'rnr', df['rnr'])

        if args.keys:
            for key in args.keys:
                if key in df.columns:
                    continue
                df.insert(len(df.columns), key, '')

        if 'labelIds' in df.columns and 'colorCode' in df.columns:
            color_fill = PatternFill(start_color="00000000", end_color="00000000")
            for i, row in df.iterrows():
                label_ids = row['labelIds']
                if label_ids:
                    color_code = row['colorCode']
                    if color_code:
                        df.at[i, 'colorCode'] = ""
                        df.at[i, 'labelIds'] = ""
                        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                        df.at[i, 'colorCode'] = ""
                        df.at[i, 'labelIds'] = ""
                        fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
                    else:
                        df.at[i, 'labelIds'] = ", ".join(label_ids)

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Vehicles"

        if args.colored:
            green_fill = PatternFill(start_color="007500", end_color="007500", fill_type="solid")
            orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
            red_fill = PatternFill(start_color="B30000", end_color="B30000", fill_type="solid")
            today = datetime.today()
            for i, row in df.iterrows():
                hu_date = datetime.strptime(row['hu'], '%Y-%m-%d')
                diff = (today - hu_date).days // 30
                fill = None
                if diff <= 3:
                    fill = green_fill
                elif diff <= 12:
                    fill = orange_fill
                else:
                    fill = red_fill
                if fill:
                    for j, _ in enumerate(row):
                        sheet.cell(row=i + 2, column=j + 2).fill = fill

        for i, row in df.iterrows():
            for j, value in enumerate(row):
                sheet.cell(row=i + 2, column=j + 2).value = value

        current_date = datetime.now().isoformat(timespec='seconds')
        output_file = f'vehicles_{current_date}.xlsx'
        with BytesIO() as excel_buffer:
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            with open(output_file, 'wb') as f:
                f.write(excel_buffer.read())

        print(f'Excel file "{output_file}" generated successfully.')
    else:
        print('Error: Failed to retrieve data from the server.')
else:
    print('Error: Failed to log in and obtain the access token.')






